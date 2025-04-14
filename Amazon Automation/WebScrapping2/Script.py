import os
import time
import sys
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
import pyinputplus as pyip
from datetime import datetime
import win32com.client as win32  # Para enviar correos
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell


# Función para cargar la plantilla HTML
def load_html_template(template_path, **kwargs):
    # Si el script se está ejecutando desde un archivo empaquetado
    if hasattr(sys, '_MEIPASS'):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))

    # Unir la ruta base con el nombre del archivo
    full_path = os.path.join(base_path, template_path)

    # Cargar el archivo HTML
    with open(full_path, 'r', encoding='utf-8') as file:
        html_content = file.read()

    return html_content.format(**kwargs)


# Función para preparar y mostrar el correo en Outlook
def prepare_email_outlook(subject, df, to_recipients, attachments):
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    
    # Get the current date and time
    current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    subject_final= f"{subject}"
    
    # Generate the email body
    email_body_content = generate_email_body(df)
    
    # Load the template and insert the body content
    script_dir = os.path.dirname(os.path.realpath(__file__))  # Path where the script is located
    template_path = os.path.join(script_dir, "ct_ob_email.html")  # Path to your email template
    
    html_template = load_html_template(template_path, body=email_body_content, site="all", report_name="Time Travel Map Compliance Report")
    
    mail.Subject = subject_final
    mail.HTMLBody = html_template
    mail.To = to_recipients
    mail.CC = ""

    # Attach files to the email
    for attachment in attachments:
        mail.Attachments.Add(attachment)

    # Use the correct Outlook account
    From = None
    for myEmailAddress in outlook.Session.Accounts:
        if "****@amazon.es" in str(myEmailAddress):
            From = myEmailAddress
            break

    if From is not None:
        mail._oleobj_.Invoke(*(64209, 0, 8, 0, From))
        mail.Display()  # Display the email for review before sending





# Función para generar el resumen en tabla de los sitios que no cumplen
# Function to generate the summary with centered symbols instead of Yes/No
def generate_non_compliance_summary(df):
    # Filter to include only the 'Site' and 'Meets_Conditions' columns
    summary_df = df[['Site', 'Meets_Conditions']]

    # If all sites meet the conditions, return a message
    if summary_df['Meets_Conditions'].all():
        return "<p>All sites meet the conditions.</p>"

    # Generate the HTML table with centered symbols
    table_html = """
    <table border="1" cellpadding="5" cellspacing="0" style="border-collapse: collapse; width: auto;">
        <thead>
            <tr>
                <th>Site</th>
                <th>Compliance</th>
            </tr>
        </thead>
        <tbody>
    """
    
    for _, row in summary_df.iterrows():
        meets_conditions = '✅' if row['Meets_Conditions'] else '❌'
        table_html += f"""
        <tr>
            <td style="text-align: center;">{row['Site']}</td>
            <td style="text-align: center;">{meets_conditions}</td>
        </tr>
        """
    
    table_html += "</tbody></table>"
    
    return table_html



def generate_email_body(df):
    
    summary_message = """
    <p>Below is a summary indicating whether each site meets all the conditions:</p>
    """
    
    # Generate the HTML table with Site and Meets_Conditions
    table_html = generate_non_compliance_summary(df)
    
    # Explanation of the conditions
    conditions_message = """
    <p>This report verifies whether the following conditions are met by each site:</p>
    <ul>
        <li>The values for 'DEFAULT:8', 'LIBRARY:8', 'LIBRARY-DEEP:8', 'PALLET-SINGLE:4', 'PALLET-DOUBLE:2', 'CASE-FLOW:3' match the expected values.</li>
        <li>The Percentile is exactly 45 ('Percentile: 45').</li>
        <li>The 'Calculate mod-to-mod' field is set to True.</li>
        <li>The month in 'Map Built Date' matches the current month.</li>
        <li>The number of days of data is greater than or equal to the expected days for the month.</li>
    </ul>
    """

    # Add the Days of Data table
    days_of_data_table = """
    <table border="1" cellpadding="5" cellspacing="0" style="border-collapse: collapse; width: auto;">
        <thead>
            <tr>
                <th>Update</th>
                <th>From</th>
                <th>To</th>
                <th>Days</th>
            </tr>
        </thead>
        <tbody>
            <tr><td>Jan / Feb</td><td>DO Not Update</td><td></td><td>-</td></tr>
            <tr><td>1st March</td><td>15th Jan</td><td>29th Feb</td><td>45</td></tr>
            <tr><td>1st April</td><td>15th Jan</td><td>31st Mar</td><td>76</td></tr>
            <tr><td>1st May</td><td>15th Jan</td><td>30th Apr</td><td>106</td></tr>
            <tr><td>1st June</td><td>15th Jan</td><td>31st May</td><td>137</td></tr>
            <tr><td>1st July</td><td>15th Jan</td><td>30th June</td><td>167</td></tr>
            <tr><td>1st Aug</td><td>15th Jan</td><td>31st July</td><td>198</td></tr>
            <tr><td>1st Sept</td><td>15th Jan</td><td>31st Aug</td><td>229</td></tr>
            <tr><td>1st Oct</td><td>15th Jan</td><td>30th Sept</td><td>259</td></tr>
            <tr><td>1st Nov</td><td>15th Jan</td><td>31st Oct</td><td>290</td></tr>
            <tr><td>Dec</td><td>DO Not Update</td><td></td><td>-</td></tr>
        </tbody>
    </table>
    """

    # Combine all the sections into one body
    full_body = summary_message + table_html + conditions_message + days_of_data_table
    return full_body





# Configuración del WebDriver
def web_driver(headless: bool = False, log_lv: int = 3, user_data_dir: str = None):
    options = Options()
    options.add_argument(f'log-level={log_lv}')
    options.add_argument("--no-sandbox")
    
    if user_data_dir:
        usrdir = user_data_dir if user_data_dir != 'auto' else os.path.join(os.getenv('LOCALAPPDATA'), 'Google', 'Chrome', 'User Data')
        options.add_argument(f'user-data-dir={usrdir}')
        options.add_argument('--profile-directory=Default')
        options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36")

    if headless:
        options.add_argument("--headless=new")
        options.add_argument("window-size=1920,1080")
    
    return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# Autenticación en Midway
def midway(driver: webdriver, pin: str, url: str = "******"):
    driver.get(url)
    otp = input("Touch Yubikey: ")
    print("Autenticación: Página cargada.")
    try:
        user_field = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.ID, "user_name"))
        )
        user_field.clear()
        user_field.send_keys(os.getlogin())
        print(f"Usuario '{os.getlogin()}' ingresado.")
        
        password_field = driver.find_element(By.ID, "password")
        password_field.send_keys(pin)
        print("PIN ingresado.")
        
        otp_field = driver.find_element(By.ID, "otp")
        otp_field.send_keys(otp)
        print("OTP ingresado.")
        
        driver.find_element(By.NAME, "commit").click()
        print("Autenticación completada.")
    except Exception as e:
        print(f"Error durante la autenticación: {e}")
        driver.quit()
        raise

# Función para intentar localizar el recuadro con varios intentos de refresh
def locate_element_with_refresh(driver, xpath, max_attempts=6):
    for attempt in range(max_attempts):
        try:
            element = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            print(f"Información disponible, elemento de referencia '{xpath}' localizado.")
            return element
        except Exception:
            if attempt < max_attempts - 1:
                print(f"Intento {attempt + 1} fallido. Recargando la página...")
                driver.refresh()
                time.sleep(15)  # Esperar 15 segundos después de cada refresh
            else:
                print(f"Elemento '{xpath}' no se pudo localizar después de {max_attempts} intentos.")
                raise

# Función para extraer información desde un conjunto de XPaths
def extract_information_by_xpath(driver, xpaths):
    extracted_data = {}
    try:
        for label, xpath in xpaths.items():
            try:
                element = WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, xpath))
                )
                extracted_data[label] = element.text.strip() if element.text else ""
                print(f"{label}: {extracted_data[label]}")
            except Exception as e:
                print(f"Error al extraer '{label}': {e}")
                extracted_data[label] = ""
    except Exception as e:
        print(f"Error al extraer la información: {e}")
        extracted_data["Error"] = "Error en la extracción de datos"
    
    return extracted_data

# Función para extraer toda la información dinámica de los Bin Type Values
def extract_dynamic_bin_type_values(driver):
    try:
        base_xpath = '//*[@id="awsui-expandable-section-0"]/span/awsui-column-layout/div/span/div/div[4]/div/div[2]'
        elements = WebDriverWait(driver, 5).until(
            EC.presence_of_all_elements_located((By.XPATH, base_xpath + '/div/div'))
        )
        values = [element.text.strip() for element in elements if element.text.strip() != "-"]
        if values:
            bin_type_values = "; ".join(values)
            print(f"Bin Type Values: {bin_type_values}")
        else:
            bin_type_values = "-"
            print("No hay información disponible en Bin Type Values.")
        return bin_type_values
    except Exception as e:
        print(f"No hay información disponible en Bin Type Values.")
        return ""

# Función para verificar las condiciones en los datos extraídos
def check_conditions(df):
    # Obtener el mes actual en formato "MM"
    current_month = datetime.now().strftime("%m")
    
    # Mapeo de Bin Type Values sin espacios después de los dos puntos
    bin_values_to_check = {
        "DEFAULT:8": "DEFAULT:8",
        "LIBRARY:8": "LIBRARY:8",
        "LIBRARY-DEEP:8": "LIBRARY-DEEP:8",
        "PALLET-SINGLE:4": "PALLET-SINGLE:4",
        "PALLET-DOUBLE:2": "PALLET-DOUBLE:2",
        "CASE-FLOW:3": "CASE-FLOW:3"
    }

    # Verificar condiciones de los Bin Type Values
    for column_label, expected_value in bin_values_to_check.items():
        df[column_label] = df["Bin Type Values"].apply(lambda x: expected_value in x if isinstance(x, str) else False)

    # Verificar el resto de las condiciones
    df['Percentile: 45'] = df['Percentile'].apply(lambda x: x.strip() == '45' if isinstance(x, str) else False)
    df['Calculate mod-to-mod:True'] = df['Calculate mod-to-mod'].apply(lambda x: x.strip().lower() == 'true' if isinstance(x, str) else False)

    # Verificar que Map Built Date sea el día 01 de cada mes
    df['Build file date check'] = df['Map Built Date'].apply(
        lambda x: (x.split('-')[1] == current_month) if '-' in x else False
    )

    # Comprobar "Days of Data" según el mes actual
    month_days_map = {
        "03": 44, "04": 76, "05": 106, "06": 137,
        "07": 167, "08": 198, "09": 229, "10": 259, "11": 290
    }
    
    # Obtener los días de datos esperados para el mes actual
    expected_days_of_data = month_days_map.get(current_month, 0)
    
    # Comprobar si los días de datos en cada fila son mayores o iguales a los esperados y mostrar la información en la columna
    def days_of_data_check(row):
        try:
            days_of_data = int(row['Days of Data'])
            if days_of_data >= expected_days_of_data:
                return f"{days_of_data} days (Check Passed, expected >= {expected_days_of_data} days)"
            else:
                return f"{days_of_data} days (Check Failed, expected >= {expected_days_of_data} days)"
        except (ValueError, TypeError):
            return f"Invalid Data (expected >= {expected_days_of_data} days)"
    
    # Aplicar la comprobación en la columna 'Days of Data Check'
    df['Days of Data Check'] = df.apply(days_of_data_check, axis=1)

    # Añadir una nueva columna con la fecha de ejecución del script
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    df['Execution Date'] = current_date
    
    # Añadir columna que indica si todas las condiciones se cumplen
    condition_columns = ['Percentile: 45', 'Calculate mod-to-mod:True', 'Build file date check'] + list(bin_values_to_check.keys()) + ['Days of Data Check']
    df['Meets_Conditions'] = df[condition_columns].all(axis=1)
    
    # Reordenar las columnas en el orden solicitado
    column_order = [
        'Execution Date', 'Site', 'Map Built Date', 'Percentile', 'Days of Data', 'End Date', 
        'Calculate mod-to-mod', 'Bin Type Values', 'DEFAULT:8', 'LIBRARY:8', 'LIBRARY-DEEP:8', 
        'PALLET-SINGLE:4', 'PALLET-DOUBLE:2', 'CASE-FLOW:3', 'Percentile: 45', 'Calculate mod-to-mod:True', 
        'Build file date check', 'Days of Data Check', 'Meets_Conditions'
    ]
    
    df = df[column_order]

    return df


# Función principal para procesar varios sitios
def process_sites(driver, sites):
    all_data = []
    for site in sites:
        print(f"Procesando {site}...")
        url = f"https://travel-time-map.eu.picking.aft.a2z.com/fc/{site}"
        driver.get(url)
        try:
            percentile_xpath = "//*[@id='awsui-expandable-section-0']/span/awsui-column-layout/div/span/div/div[1]/div/div[3]/div[2]"
            locate_element_with_refresh(driver, percentile_xpath)
        except Exception as e:
            print(f"Error al cargar la página para {site}: {e}")
            all_data.append({"Site": site, "Error": "No data available"})
            continue
        
        try:
            xpaths = {
                "Map Built Date": "//*[@id='awsui-expandable-section-0']/span/awsui-column-layout/div/span/div/div[1]/div/div[1]/div[2]",
                "Percentile": "//*[@id='awsui-expandable-section-0']/span/awsui-column-layout/div/span/div/div[1]/div/div[3]/div[2]",
                "Days of Data": "//*[@id='awsui-expandable-section-0']/span/awsui-column-layout/div/span/div/div[2]/div/div[1]/div[2]",
                "End Date": "//*[@id='awsui-expandable-section-0']/span/awsui-column-layout/div/span/div/div[2]/div/div[2]/div[2]",
                "Calculate mod-to-mod": "//*[@id='awsui-expandable-section-0']/span/awsui-column-layout/div/span/div/div[2]/div/div[3]/div[2]"
            }
            extracted_info = extract_information_by_xpath(driver, xpaths)
            bin_type_values = extract_dynamic_bin_type_values(driver)
            extracted_info["Bin Type Values"] = bin_type_values
            extracted_info["Site"] = site
            all_data.append(extracted_info)
        except Exception as e:
            print(f"Error procesando {site}: {e}")
            all_data.append({"Site": site, "Error": "No data available"})
    return all_data



def format_excel_with_groups(output_file):
    # Cargar el archivo de Excel generado
    wb = load_workbook(output_file)
    ws = wb.active

    # Definir colores de relleno
    fill_group_1 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Gris
    fill_group_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Blanco
    fill_true = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")    # Verde para True
    fill_false = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")   # Rojo para False

    # Definir columnas para cada grupo
    group_1_columns = ['Execution Date', 'Site', 'Map Built Date', 'Percentile', 'Days of Data', 'End Date', 'Calculate mod-to-mod', 'Bin Type Values']
    group_2_columns = ['DEFAULT:8', 'LIBRARY:8', 'LIBRARY-DEEP:8', 'PALLET-SINGLE:4', 'PALLET-DOUBLE:2', 'CASE-FLOW:3', 'Percentile: 45', 'Calculate mod-to-mod:True', 'Build file date check', 'Days of Data Check', 'Meets_Conditions']

    # Insertar una fila extra en la parte superior para las celdas fusionadas de los grupos
    ws.insert_rows(1)

    # Obtener las letras de las columnas para cada grupo
    group_1_indexes = [ws[2].index(cell) + 1 for cell in ws[2] if cell.value in group_1_columns]
    group_2_indexes = [ws[2].index(cell) + 1 for cell in ws[2] if cell.value in group_2_columns]

    # Fusionar las celdas para los grupos en la nueva fila agregada
    ws.merge_cells(start_row=1, start_column=group_1_indexes[0], end_row=1, end_column=group_1_indexes[-1])
    ws.merge_cells(start_row=1, start_column=group_2_indexes[0], end_row=1, end_column=group_2_indexes[-1])

    # Añadir el título del grupo "Extracted Data"
    cell = ws.cell(row=1, column=group_1_indexes[0])
    cell.value = "Extracted Data"
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    # Añadir el título del grupo "Checked Data"
    cell = ws.cell(row=1, column=group_2_indexes[0])
    cell.value = "Checked Data"
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    # Aplicar colores a las celdas en los títulos y los datos
    for row in ws.iter_rows(min_row=2):  # Comienza desde la fila de datos (segunda fila)
        for idx in group_1_indexes:
            row[idx - 1].fill = fill_group_1  # Aplicar color al grupo 1
        for idx in group_2_indexes:
            row[idx - 1].fill = fill_group_2  # Aplicar color al grupo 2

    # Aplicar colores condicionales a las columnas con condiciones verificadas (True o False)
    for row in ws.iter_rows(min_row=3):  # Desde la fila 3 (los datos) en adelante
        for idx in group_2_indexes:
            cell_value = row[idx - 1].value
            if cell_value == True:  # Si es True, verde
                row[idx - 1].fill = fill_true
            elif cell_value == False:  # Si es False, rojo
                row[idx - 1].fill = fill_false

    # Buscar la columna "Days of Data Check" y aplicar colores
    days_of_data_col_index = group_2_indexes[group_2_columns.index('Days of Data Check')]  # Obtener el índice de la columna 'Days of Data Check'
    for row in ws.iter_rows(min_row=3):  # Desde la fila 3 (los datos) en adelante
        cell = row[days_of_data_col_index - 1]
        if cell.value and "Check Failed" in str(cell.value):  # Si contiene "Check Failed", aplicar el color rojo
            cell.fill = fill_false
        elif cell.value and "Check Passed" in str(cell.value):  # Si contiene "Check Passed", aplicar el color verde
            cell.fill = fill_true

    # Ajustar el ancho de las columnas basado en el contenido (excepto las celdas fusionadas)
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)  # Obtener la letra de la columna

        for cell in col:
            if not isinstance(cell, MergedCell):  # Ignorar las celdas fusionadas
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))  # Obtener el valor más largo
                except:
                    pass
        adjusted_width = (max_length + 2)  # Ajustar el ancho, añadiendo algo de espacio extra
        ws.column_dimensions[column_letter].width = adjusted_width

    # Guardar el archivo con el formato actualizado
    wb.save(output_file)




# Función principal para procesar y generar el Excel con formato y grupos
def generate_excel_with_format_and_groups(df, output_file):
    # Guardar DataFrame en archivo Excel sin formato
    df.to_excel(output_file, index=False)

    # Aplicar formato al archivo Excel
    format_excel_with_groups(output_file)
    print(f"Archivo Excel guardado y formateado en: {output_file}")



# Script principal
if __name__ == "__main__":
    # Registrar el tiempo de inicio
    start_time = time.time()

    headless_mode = True
    driver = web_driver(headless_mode, user_data_dir='auto')
    try:
        pin = pyip.inputPassword("Enter your PIN > ")
        midway(driver=driver, pin=pin)
    except Exception as e:
        print(f"Fallo en la autenticación: {e}")
        exit(1)

    #sites = ["435g", "aY542", "6543Y", "54367Y", "64356>Y", "63456Y", "643y"]
    sites = ["Y6534"]
    extracted_data = process_sites(driver, sites)
    df = pd.DataFrame(extracted_data)

    # Verificar si DataFrame no está vacío
    if not df.empty:
        df = check_conditions(df)
    else:
        print("No se extrajo ninguna información de los sitios.")

    # Guardar el archivo en el mismo directorio donde se encuentra el script
    script_dir = os.path.dirname(os.path.realpath(__file__))  # Obtiene la ruta del script
    output_file = os.path.join(script_dir, "TTM_checked.xlsx")

    # Generar Excel y aplicar formato
    generate_excel_with_format_and_groups(df, output_file)
    
    # try:
        # df.to_excel(output_file, index=False)
        #print(f"Información guardada en {output_file}")
    # except Exception as e:
        # print(f"Error al guardar el archivo Excel: {e}")


    # Generar el resumen en tabla de los sitios non-compliance
    non_compliance_summary = generate_non_compliance_summary(df)
    # Cargar la plantilla HTML
    template_path = "ct_ob_email.html"  # Nombre del archivo sin la ruta completa
    html_body = load_html_template(template_path, body=non_compliance_summary, site="all", report_name="Time Travel Map Compliance Report")

    # Enviar correo con el archivo adjunto
    print("Generando y mostrando el email con el archivo adjunto...")
    current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S") # Obtener la fecha y hora actuales
    msg_subjet = f"Time Travel Map Compliance Report"
    subject = f"{msg_subjet} - Execution Time: {current_datetime}" # Incluir la fecha y hora en el subject
    to_recipients = ""  # Completar al enviar
    attachments = [output_file]



    # Preparar y mostrar el email
    prepare_email_outlook(subject, df, to_recipients, attachments)


    # Calcular el tiempo total transcurrido
    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f"Tiempo total de ejecución: {elapsed_time:.2f} segundos")

    # print("Manteniendo la ventana abierta durante 15 segundos...")
    # time.sleep(15)
    driver.quit()
